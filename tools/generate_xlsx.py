from openpyxl import load_workbook
import datetime
import shutil
import os

def generate_audit_xlsx(worker_name: str):
    worker_xlsx_path: str = f'z:\\{worker_name}\\报告审核记录表-{worker_name}-{today_date}.xlsx'
    shutil.copyfile(template_path, worker_xlsx_path)
    record_wb = load_workbook(worker_xlsx_path)
    record_sht = record_wb['Sheet1']
    report_info_list: list = []

    for file_name in os.listdir(f'z:\\{worker_name}'):
        if not os.path.splitext(file_name)[1] == '.xlsx':
            report_name: str = file_name.replace('.pptx', '').replace('.ppt', '').replace('R','报告')
            temp_list: list = report_name.split('_')
            temp_list[3] = str(datetime.datetime.strptime(temp_list[3], "%y%m%d"))[:10]
            report_info_list.append(temp_list)
    # print('已经全部载入报告信息！')

    for i in range(2, 2 + len(report_info_list)):
        for j in range(2, 6):
            record_sht.cell(i, j).value = report_info_list[i-2][j-2]
        record_sht.cell(i, 7).value = today_date_2
        record_sht.cell(i, 10).value = worker_name

    record_wb.save(worker_xlsx_path)
    del record_wb
    print(f'[OK] --> 已生成 {worker_name} 的审核反馈表！')


def generate_folder(worker_name: str):
    os.makedirs(os.path.join('Z:', worker_name, '合格报告'))
    os.makedirs(os.path.join('Z:', worker_name, '不合格报告'))


today_date: str = str(datetime.date.today()).replace("-", "").replace("2019", "19")
today_date_2: str = str(datetime.date.today())
template_path: str = 'G:\\报告审核记录表.xlsx'

worker_name_list: list = ['赵康明', '马云超', '张喆', '郭点点', '郑聪颖', '彭家宜', '李海玉', '赵文慧', '王潼恩', '龚新潮', '李群', '翟银凤']

for worker_name in worker_name_list:
    if os.path.exists(os.path.join('Z:', worker_name)):
        generate_audit_xlsx(worker_name)
        generate_folder(worker_name)

print('\n' + '-'*40 + ' 已经全部生成 ' + '-'*40)

